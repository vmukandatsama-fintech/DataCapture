from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.http import HttpResponse
from unfold.admin import ModelAdmin as UnfoldModelAdmin
import csv
import io
import re

from .models import User, Grower, Floor, RejectionClassification, RejectedBale, ReleaseFromHold
from .forms import UserImportForm, GrowerImportForm, AdminUserCreationForm


@admin.register(User)
class CustomUserAdmin(UnfoldModelAdmin, UserAdmin):
    model = User
    add_form = AdminUserCreationForm

    add_fieldsets = (
        (None, {
            "fields": ("username", "role", "pin", "pin_confirm", "is_staff", "must_change_pin"),
        }),
    )

    fieldsets = UserAdmin.fieldsets + (
        ("DataCapture Fields", {
            "fields": ("role", "pin", "must_change_pin"),
        }),
    )

    list_display = ("username", "role", "must_change_pin", "is_staff")

    def save_model(self, request, obj, form, change):
        if not change and isinstance(form, AdminUserCreationForm):
            # PIN already hashed by AdminUserCreationForm.save(); bypass UserAdmin password logic
            obj.save()
        else:
            super().save_model(request, obj, form, change)

    # URL for import
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-users/",
                self.admin_site.admin_view(self.import_users),
                name="accounts_user_import",
            ),
        ]
        return custom_urls + urls

    # Import logic
    def import_users(self, request):
        if request.method == "POST":
            form = UserImportForm(request.POST, request.FILES)

            if form.is_valid():
                file = form.cleaned_data["csv_file"]
                decoded = file.read().decode("utf-8").splitlines()
                reader = csv.DictReader(decoded)

                count = 0

                for row in reader:
                    if not row.get("username"):
                        continue

                    User.objects.create(
                        username=row["username"],
                        role=row.get("role", "capturer"),
                        pin=make_password(row.get("pin", "1234")),
                        must_change_pin=True,
                        is_staff=True,
                    )
                    count += 1

                messages.success(request, f"{count} users imported successfully!")
                return redirect("../")

        else:
            form = UserImportForm()

        return render(request, "admin/import_users.html", {"form": form})

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context["import_url"] = "import-users/"
        return super().changelist_view(request, extra_context=extra_context)


# ── Grower column header normaliser ──────────────────────────────────────────

def _norm(s):
    return re.sub(r'[\s_\-#]+', '', s).lower()


_SENTENCE_CASE_FIELDS = {
    'grower_name', 'timb_name', 'master_id', 'timb_id',
    'area', 'group_name', 'manager', 'area_manager', 'ltech',
    'stop_order_number', 'contract_status', 'biometrics',
}


# Maps normalised CSV header → (model_field, type_converter)
_GROWER_COLUMN_MAP = {
    'growerid':          ('grower_id',            str),
    'growerno':          ('grower_id',            str),
    'growername':        ('grower_name',          str),
    'timbname':          ('timb_name',            str),
    'masterid':          ('master_id',            str),
    'timbid':            ('timb_id',              str),
    'scheme':            ('scheme',               str),
    'cy26ha':            ('cy26_ha',              lambda v: float(v) if v else None),
    'cy26hectares':      ('cy26_ha',              lambda v: float(v) if v else None),
    'hectares':          ('cy26_ha',              lambda v: float(v) if v else None),
    'ha':                ('cy26_ha',              lambda v: float(v) if v else None),
    'stopordernumber':   ('stop_order_number',    str),
    'stoporder':         ('stop_order_number',    str),
    'contractstatus':    ('contract_status',      str),
    'status':            ('contract_status',      str),
    'areacode':          ('area_code',            int),
    'area':              ('area',                 str),
    'groupcode':         ('group_code',           int),
    'groupname':         ('group_name',           str),
    'group':             ('group_name',           str),
    'managercode':       ('manager_code',         int),
    'manager':           ('manager',              str),
    'areamanagercode':   ('area_manager_code',    int),
    'areamanager':       ('area_manager',         str),
    'am':                ('area_manager',         str),
    'amcode':            ('area_manager_code',    int),
    'ltechcode':              ('ltech_code', int),
    'leadtechcode':           ('ltech_code', int),
    'leaftechcode':           ('ltech_code', int),
    'leaftechniciancode':     ('ltech_code', int),
    'ltech':                  ('ltech',      str),
    'leadtechnician':         ('ltech',      str),
    'leaftechnician':         ('ltech',      str),
    'biometrics':             ('biometrics', str),
    'timbstopordervalue':     ('timb_stop_order_value', lambda v: float(v) if v else None),
    'stopordervalue':    ('timb_stop_order_value', lambda v: float(v) if v else None),
}


@admin.register(Grower)
class GrowerAdmin(UnfoldModelAdmin):
    list_display = (
        'grower_id', 'grower_name', 'scheme', 'area',
        'group_name', 'manager', 'cy26_ha', 'contract_status', 'biometrics',
    )
    list_filter = ('scheme', 'area', 'group_name', 'contract_status', 'biometrics')
    search_fields = ('grower_id', 'grower_name', 'timb_name', 'timb_id')
    ordering = ('grower_id',)
    list_per_page = 50

    fieldsets = (
        ('Identity', {
            'fields': ('grower_id', 'grower_name', 'timb_name', 'master_id', 'timb_id'),
        }),
        ('Classification', {
            'fields': ('scheme', 'cy26_ha', 'stop_order_number', 'contract_status', 'biometrics'),
        }),
        ('Area & Group', {
            'fields': ('area_code', 'area', 'group_code', 'group_name'),
        }),
        ('Personnel', {
            'fields': (
                'manager_code', 'manager',
                'area_manager_code', 'area_manager',
                'ltech_code', 'ltech',
            ),
        }),
        ('TIMB Financial', {
            'fields': ('timb_stop_order_value',),
        }),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'import-growers/',
                self.admin_site.admin_view(self.import_growers),
                name='accounts_grower_import',
            ),
            path(
                'import-growers/sample.csv',
                self.admin_site.admin_view(self.sample_csv),
                name='accounts_grower_sample_csv',
            ),
        ]
        return custom_urls + urls

    def sample_csv(self, request):
        import io
        from django.http import HttpResponse

        headers = [
            'Grower ID', 'Grower Name', 'TIMB Name', 'Master ID', 'TIMB ID',
            'Scheme', 'CY26 Ha', 'Stop Order #', 'Contract Status',
            'Area Code', 'Area', 'Group Code', 'Group', 'Manager Code', 'Manager',
            'Area Manager Code', 'Area Manager', 'Leaf Technician Code', 'Leaf Technician',
            'TIMB Stop Order Value', 'Biometrics',
        ]
        rows = [
            [
                'V102398', 'JOHN MOYO', 'J MOYO', 'M001', 'T001',
                'A Scheme', '2.50', 'SO-001', 'Active',
                '1', 'HARARE', '10', 'HARARE NORTH', '20', 'T SMITH',
                '30', 'R JONES', '40', 'P DUBE',
                '1500.00', 'Enrolled',
            ],
            [
                'V102399', 'JANE MUTASA', '', '', '',
                'B Scheme', '1.75', '', 'Active',
                '1', 'HARARE', '10', 'HARARE NORTH', '20', 'T SMITH',
                '30', 'R JONES', '40', 'P DUBE',
                '', 'Not Enrolled',
            ],
        ]

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)

        response = HttpResponse(buf.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="grower_import_sample.csv"'
        return response

    def import_growers(self, request):
        if request.method == 'POST':
            form = GrowerImportForm(request.POST, request.FILES)
            if form.is_valid():
                file = form.cleaned_data['csv_file']
                try:
                    decoded = file.read().decode('utf-8-sig').splitlines()
                except UnicodeDecodeError:
                    file.seek(0)
                    decoded = file.read().decode('latin-1').splitlines()

                reader = csv.DictReader(decoded)
                headers = reader.fieldnames or []
                col_map = {}
                for h in headers:
                    key = _norm(h)
                    if key in _GROWER_COLUMN_MAP:
                        col_map[h] = _GROWER_COLUMN_MAP[key]

                created = updated = skipped = 0
                errors = []

                with transaction.atomic():
                    for i, row in enumerate(reader, start=2):
                        # Resolve grower_id
                        grower_id = None
                        for h, (field, conv) in col_map.items():
                            if field == 'grower_id':
                                raw = row.get(h, '').strip()
                                grower_id = raw if raw else None
                                break

                        if not grower_id:
                            skipped += 1
                            continue

                        defaults = {}
                        for h, (field, conv) in col_map.items():
                            if field == 'grower_id':
                                continue
                            raw = row.get(h, '').strip()
                            try:
                                value = conv(raw) if raw else None
                                if isinstance(value, str) and field in _SENTENCE_CASE_FIELDS:
                                    value = value.title()
                                defaults[field] = value
                            except (ValueError, TypeError) as e:
                                errors.append(f"Row {i}: {field} — {e}")

                        # Required integer fields default to 0 if missing
                        for int_field in ('area_code', 'group_code', 'manager_code',
                                          'area_manager_code', 'ltech_code'):
                            if defaults.get(int_field) is None:
                                defaults[int_field] = 0

                        # Required string fields default to empty string if missing
                        for str_field in ('area', 'group_name', 'manager',
                                          'area_manager', 'ltech', 'scheme'):
                            if defaults.get(str_field) is None:
                                defaults[str_field] = ''

                        _, was_created = Grower.objects.update_or_create(
                            grower_id=grower_id,
                            defaults=defaults,
                        )
                        if was_created:
                            created += 1
                        else:
                            updated += 1

                summary = f"{created} created, {updated} updated, {skipped} skipped."
                if errors:
                    messages.warning(request, f"Import complete with issues: {summary} Errors: {'; '.join(errors[:5])}")
                else:
                    messages.success(request, f"Import successful — {summary}")
                return redirect('../')
        else:
            form = GrowerImportForm()

        column_info = [
            {'headers': 'Grower ID, Grower No',             'field': 'grower_id',            'required': 'Yes'},
            {'headers': 'Grower Name',                       'field': 'grower_name',          'required': 'Yes'},
            {'headers': 'Scheme',                            'field': 'scheme',               'required': 'Yes'},
            {'headers': 'Area Code',                         'field': 'area_code',            'required': 'Yes'},
            {'headers': 'Area',                              'field': 'area',                 'required': 'Yes'},
            {'headers': 'Group Code',                        'field': 'group_code',           'required': 'Yes'},
            {'headers': 'Group, Group Name',                 'field': 'group_name',           'required': 'Yes'},
            {'headers': 'Manager Code',                      'field': 'manager_code',         'required': 'Yes'},
            {'headers': 'Manager',                           'field': 'manager',              'required': 'Yes'},
            {'headers': 'Area Manager Code, AM Code',        'field': 'area_manager_code',    'required': 'Yes'},
            {'headers': 'Area Manager, AM',                  'field': 'area_manager',         'required': 'Yes'},
            {'headers': 'Leaf Technician Code, LTech Code',   'field': 'ltech_code',           'required': 'Yes'},
            {'headers': 'Leaf Technician, LTech',            'field': 'ltech',                'required': 'Yes'},
            {'headers': 'CY26 Ha, CY26 Hectares, Ha',        'field': 'cy26_ha',              'required': 'No'},
            {'headers': 'TIMB Name',                         'field': 'timb_name',            'required': 'No'},
            {'headers': 'TIMB ID',                           'field': 'timb_id',              'required': 'No'},
            {'headers': 'Master ID',                         'field': 'master_id',            'required': 'No'},
            {'headers': 'Stop Order #, Stop Order Number',   'field': 'stop_order_number',    'required': 'No'},
            {'headers': 'Contract Status, Status',           'field': 'contract_status',      'required': 'No'},
            {'headers': 'TIMB Stop Order Value',             'field': 'timb_stop_order_value','required': 'No'},
            {'headers': 'Biometrics',                        'field': 'biometrics',           'required': 'No'},
        ]
        return render(request, 'admin/import_growers.html', {'form': form, 'column_info': column_info})

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['import_url'] = 'import-growers/'
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(Floor)
class FloorAdmin(UnfoldModelAdmin):
    list_display = ('name', 'location')
    search_fields = ('name', 'location')
    ordering = ('name',)


@admin.register(RejectionClassification)
class RejectionClassificationAdmin(UnfoldModelAdmin):
    list_display = ('code', 'description')
    search_fields = ('code', 'description')
    ordering = ('code',)


@admin.register(ReleaseFromHold)
class ReleaseFromHoldAdmin(UnfoldModelAdmin):
    list_display = ('resolution_date', 'ticket_number', 'resolution', 'reference', 'floor', 'created_by')
    list_filter = ('resolution_date', 'resolution', 'floor')
    search_fields = ('ticket_number', 'reference')
    ordering = ('-resolution_date',)
    date_hierarchy = 'resolution_date'


@admin.register(RejectedBale)
class RejectedBaleAdmin(UnfoldModelAdmin):
    list_display = ('date', 'floor', 'grower_number', 'grower_name', 'ticket_number', 'lot_number', 'group_number', 'classification', 'created_by')
    list_filter = ('date', 'floor', 'classification')
    search_fields = ('grower_number', 'grower_name', 'ticket_number')
    ordering = ('-date',)
    date_hierarchy = 'date'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'lifecycle-report/',
                self.admin_site.admin_view(self.lifecycle_report),
                name='reports_rejected_bales_lifecycle',
            ),
        ]
        return custom_urls + urls

    # ── Report view ───────────────────────────────────────────────────────────

    def lifecycle_report(self, request):
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        floor_id = request.GET.get('floor', '')
        resolution_filter = request.GET.get('resolution', '')
        fmt = request.GET.get('format', '')

        floors = Floor.objects.order_by('name')

        if fmt in ('csv', 'xlsx', 'preview'):
            rows = self._get_lifecycle_rows(date_from, date_to, floor_id, resolution_filter)

            if fmt == 'csv':
                return self._export_csv(rows, date_from, date_to)
            if fmt == 'xlsx':
                return self._export_xlsx(rows, date_from, date_to)

            # preview
            context = {
                'floors': floors,
                'date_from': date_from,
                'date_to': date_to,
                'floor_id': floor_id,
                'resolution_filter': resolution_filter,
                'rows': rows[:100],
                'total_count': len(rows),
                'has_results': True,
            }
        else:
            context = {
                'floors': floors,
                'date_from': date_from,
                'date_to': date_to,
                'floor_id': floor_id,
                'resolution_filter': resolution_filter,
                'has_results': False,
            }

        return render(request, 'admin/reports/rejected_bales_lifecycle.html', context)

    # ── Data helper ───────────────────────────────────────────────────────────

    def _get_lifecycle_rows(self, date_from, date_to, floor_id, resolution_filter):
        from django.db.models import Prefetch

        release_qs = ReleaseFromHold.objects.select_related('created_by').order_by('resolution_date')

        qs = RejectedBale.objects.select_related(
            'floor', 'classification', 'created_by'
        ).prefetch_related(
            Prefetch('releases', queryset=release_qs)
        ).order_by('-date', 'ticket_number')

        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if floor_id:
            qs = qs.filter(floor_id=floor_id)

        rows = []
        for bale in qs:
            base = {
                'rejection_date': bale.date,
                'floor': bale.floor.name if bale.floor else '',
                'grower_number': bale.grower_number,
                'grower_name': bale.grower_name,
                'ticket_number': bale.ticket_number,
                'lot_number': bale.lot_number,
                'group_number': bale.group_number,
                'classification': bale.classification.code if bale.classification else '',
                'ntrm_type': bale.ntrm_type or '',
                'captured_by': bale.created_by.username if bale.created_by else '',
            }
            releases = list(bale.releases.all())
            if releases:
                for rel in releases:
                    status = rel.resolution
                    if resolution_filter and status != resolution_filter:
                        continue
                    rows.append({**base,
                        'status': status,
                        'resolution_date': rel.resolution_date,
                        'reference': rel.reference or '',
                        'resolved_by': rel.created_by.username if rel.created_by else '',
                    })
            else:
                if resolution_filter and resolution_filter != 'Pending':
                    continue
                rows.append({**base,
                    'status': 'Pending',
                    'resolution_date': '',
                    'reference': '',
                    'resolved_by': '',
                })

        return rows

    # ── CSV export ────────────────────────────────────────────────────────────

    _REPORT_HEADERS = [
        'Rejection Date', 'Floor', 'Grower Number', 'Grower Name',
        'Ticket Number', 'Lot Number', 'Group Number', 'Classification',
        'NTRM Type', 'Captured By', 'Status', 'Resolution Date',
        'Reference', 'Resolved By',
    ]

    @staticmethod
    def _row_values(row):
        return [
            row['rejection_date'], row['floor'], row['grower_number'],
            row['grower_name'], row['ticket_number'], row['lot_number'],
            row['group_number'], row['classification'], row['ntrm_type'],
            row['captured_by'], row['status'], row['resolution_date'],
            row['reference'], row['resolved_by'],
        ]

    def _export_csv(self, rows, date_from, date_to):
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(self._REPORT_HEADERS)
        for row in rows:
            writer.writerow(self._row_values(row))

        suffix = f"{date_from or 'all'}_{date_to or 'all'}"
        response = HttpResponse(buf.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="rejected_bales_lifecycle_{suffix}.csv"'
        return response

    # ── Excel export ──────────────────────────────────────────────────────────

    def _export_xlsx(self, rows, date_from, date_to):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Lifecycle'

        header_fill = PatternFill(start_color='157D3D', end_color='157D3D', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(self._REPORT_HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        STATUS_COLORS = {
            'Pending': 'B45309',
            'Release to Grower': '166534',
            'ReOffer': '1D4ED8',
        }

        for r, row in enumerate(rows, 2):
            values = self._row_values(row)
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=c, value=str(val) if val else '')
            # Colour the Status column (col 11)
            status_cell = ws.cell(row=r, column=11)
            color = STATUS_COLORS.get(row['status'])
            if color:
                status_cell.font = Font(bold=True, color=color)

        col_widths = [14, 16, 14, 24, 16, 10, 13, 16, 18, 16, 20, 16, 26, 16]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        suffix = f"{date_from or 'all'}_{date_to or 'all'}"
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="rejected_bales_lifecycle_{suffix}.xlsx"'
        return response
